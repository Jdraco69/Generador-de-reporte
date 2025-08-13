import openpyxl
from tabulate import tabulate

excel_data = openpyxl.load_workbook("Ventas.xlsx")

dataframe = excel_data.active

data= []

for row in range(1, dataframe.max_row):
    _row=[row,]
    for col in dataframe.iter_cols(1, dataframe.max_column):
       _row.append(col[row].value)
    data.append(_row)    
  
print(tabulate(data))    